# https://engkimbs.tistory.com/625?category=762758

import requests
import numpy as  np
import pandas as pd
from bs4 import BeautifulSoup
import re
import manageFile
from openpyxl import load_workbook
import kakao
import my_email


# 자릿수 옵션
pd.options.display.float_format = '{:.2f}'.format

def analyzeStock(code):
    ## request and parse
    try:
        url = "https://finance.naver.com/item/sise.nhn?code={}".format(code)
        html = requests.get(url).text
        parser = BeautifulSoup(html,'html.parser') 
        financeInfo_html = parser.select('div.section.inner_sub')[0]

        url = "https://finance.naver.com/item/main.nhn?code={}".format(code)
        html = requests.get(url).text
        parser = BeautifulSoup(html,'html.parser') 
        finance_html = parser.select('div.section.cop_analysis div.sub_section')[0]
        name_html = parser.select('div.wrap_company')[0]
    except Exception as e:
        analyzeResult = pd.Series(['ERROR no data'], index = ['result'])
        return analyzeResult

    # 이름
    company_name = name_html.select('h2')[0].text

    # 현재가
    now_price = financeInfo_html.select('tr')[0]
    now_price = [item.get_text().strip() for item in now_price.select('td')][0]
    now_price = float(now_price.replace(',',''))

    # 시가총액
    market_value = financeInfo_html.select('tr')[11] 
    market_value = [item.get_text().strip() for item in market_value.select('td')][0]
    market_value = market_value.replace('억원','').replace(',','')

    # 분기 연간 정보 index
    finance_date = [item.get_text().strip() for item in finance_html.select('thead th')][3:13]
    if finance_date[0]=='':
        analyzeResult = pd.Series(['ERROR no data'], index = ['result'])
        return analyzeResult

    # 재무정보 명칭
    finance_index = [item.get_text().strip() for item in finance_html.select('th.h_th2')][3:]

    # 재무재표 데이터
    finance_data = [item.get_text().strip() for item in finance_html.select('td')]
    finance_data = np.array(finance_data)
    finance_data.resize(len(finance_index),10)
    finance = pd.DataFrame(data=finance_data[0:,0:], index=finance_index, columns=finance_date)
    annual_finance = finance.iloc[:, :4]
    quarter_finance = finance.iloc[:, 4:]

    ## 분석 및 평가 -------------------------------------------------------------------
    score = 100
    analyzeResult = pd.Series(['{}'.format(score)], index =['result']) # error 때문에 string 으로 햇음
    
    annual_isIncreasing = annual_finance.iloc[:,0:3].apply(lambda x: x.is_monotonic_increasing, axis=1)
    quater_isIncreasing = quarter_finance.iloc[:,0:5].apply(lambda x: x.is_monotonic_increasing, axis=1)

    # 연간 매출 감소
    if annual_isIncreasing['매출액'] == False:
        score -= 10
        warning = pd.Series(['연간 매출 감소'])
        analyzeResult = pd.concat([analyzeResult,warning])

    # 연간 영업이익 감소
    if annual_isIncreasing['영업이익'] == False:
        score -= 10
        warning = pd.Series(['연간 영업이익 감소'])
        analyzeResult = pd.concat([analyzeResult,warning])

    # 연간 부채 증가
    if annual_isIncreasing['부채비율'] == True:
        score -= 5
        warning = pd.Series(['부채비율 증가'])
        analyzeResult = pd.concat([analyzeResult,warning])

    # 연간 EPS 증가
    if annual_isIncreasing['EPS(원)'] == False:
        score -= 5
        warning = pd.Series(['연간 EPS 감소'])
        analyzeResult = pd.concat([analyzeResult,warning])

    # 연간 배당금 감소
    if annual_isIncreasing['주당배당금(원)'] == False:
        score -= 5
        warning = pd.Series(['연간 주당배당금 감소'])
        analyzeResult = pd.concat([analyzeResult,warning])

    # 분기간 매출액 감소
    if quater_isIncreasing['매출액'] == False:
        score -= 5
        warning = pd.Series(['분기간 매출액 감소'])
        analyzeResult = pd.concat([analyzeResult,warning])

    # 분기간 영업이익 감소
    if quater_isIncreasing['영업이익'] == False:
        score -= 5
        warning = pd.Series(['분기간 영업이익 감소'])
        analyzeResult = pd.concat([analyzeResult,warning])
    
    # 부채비율 100% 미만 (현명한투자자)
    if float(quarter_finance.iloc[6,4].replace(',','')) > 100:
        score -= 5
        warning = pd.Series(['부채비율 100% 이상'])
        analyzeResult = pd.concat([analyzeResult,warning])

    # 유보율 200% 이상 (그래야 약세장을 견딤)
    if float(quarter_finance.iloc[8,4].replace(',','')) < 200:
        score -= 5
        warning = pd.Series(['유보율 200% 미만'])
        analyzeResult = pd.concat([analyzeResult,warning])

    # 적당한 주가 수익비율 (현명한투자자)
    expectationPrice = 15 * (float(annual_finance.iloc[9,0].replace(',','')) + float(annual_finance.iloc[9,1].replace(',','')) +
    float(annual_finance.iloc[9,2].replace(',',''))) / 3
    if now_price > expectationPrice:
        score -= 10
        warning = pd.Series(['비싼 주가 수익비율'])
        analyzeResult = pd.concat([analyzeResult,warning])
        reason = pd.Series(['{}>{:.0f}'.format(now_price,expectationPrice)])
        analyzeResult = pd.concat([analyzeResult,reason])
    
    
    # valuation when per = 10
    try:
        profitRate = float(quarter_finance.iloc[1,4].replace(',','')) / float(quarter_finance.iloc[1,0].replace(',',''))
        expectationProfit = profitRate * (float(quarter_finance.iloc[1,1].replace(',','')) + float(quarter_finance.iloc[1,2].replace(',','')) +
        float(quarter_finance.iloc[1,3].replace(',',''))) + float(quarter_finance.iloc[1,4].replace(',',''))
        annual_finance.iloc[1,3] = '{:.0f}'.format(expectationProfit) # 예측값 입력
        if float(market_value.replace(',','')) > expectationProfit * 10:
            score -= 20
            warning = pd.Series(['PER 10 valuation 부적합'])
            analyzeResult = pd.concat([analyzeResult,warning])
            reason = pd.Series(['{}>{:.0f}'.format(market_value,expectationProfit * 10)])
            analyzeResult = pd.concat([analyzeResult,reason])

        # 예측된 연간 영업이익 감소
        if expectationProfit < float(annual_finance.iloc[1,2].replace(',','')):
            score -= 10
            warning = pd.Series(['예측된 연간 영업이익 감소'])
            analyzeResult = pd.concat([analyzeResult,warning])

    except Exception as lackofdata:
        profitRate = 0
        expectationProfit = 0 
        pass
    
    # 유효한 값들 계산
    try:
        proper_value_rate = float(expectationProfit * 10) / float(market_value) # 적정 시가총액
        proper_price = now_price * proper_value_rate # 적정 주가
        now_dividend = float(annual_finance.iloc[13,3].replace(',',''))
        average_dividend = (float(annual_finance.iloc[13,0].replace(',','')) + float(annual_finance.iloc[13,1].replace(',','')) + float(annual_finance.iloc[13,2].replace(',','')))/3
        rate_dividend = average_dividend / now_price * 100 # 배당 수익률
        target_price_by_dividend =  100 * now_dividend / 3
        target_price_by_avg_dividend =  100 * average_dividend / 3

    except Exception as lackofdata:
        now_dividend = 0
        average_dividend = 0 
        rate_dividend = 0
        target_price_by_dividend = 0
        target_price_by_avg_dividend = 0
        pass

    # 최종 스코어 저장    
    analyzeResult.iloc[0] = '{}'.format(score)

    ## 분석결과 저장
    if score > 80 and score != "ERROR no data":
        annual_isIncreasing = annual_isIncreasing.replace(True,'up').replace(False,'down')
        quater_isIncreasing = quater_isIncreasing.replace(True,'up').replace(False,'down')
        analyzeResultSheet = pd.concat([annual_finance,annual_isIncreasing,quarter_finance,quater_isIncreasing],axis=1) 
        #analyzeResultSheet = pd.concat([analyzeResultSheet,analyzeResult], axis = 1)
        
        file_dir = "C:/Users/조성현/Desktop/super_ant/recommend_list.xlsx"
        book = load_workbook(file_dir)
        writer = pd.ExcelWriter(file_dir, engine='openpyxl')
        writer.book = book

        analyzeResultSheet.to_excel(writer, sheet_name='{}'.format(company_name))
        analyzeResult.to_excel(writer, sheet_name='{}'.format(company_name),startcol=15)

        writer.save()
        writer.close()

        # # 알림 보내기
        # if isCodeExist(code,'recommendedList.csv') == False:
        #     print('메일')
        #     # 카톡보내기
        #     # kakao.send_to_kakao('새로운 종목이 감지 되었습니다. \n {}({}) {:.0f}(원) \n {}'.format(company_name,code,now_price,url))
            
        #     # 메일 보내기
        #     str_subject = '새로운 종목이 감지 되었습니다.'
            
        #     str_text = '{}({}) {:.0f}(원) \n {} \n\n'.format(company_name,code,now_price,url)
        #     reason = ', '.join(analyzeResult[1:].tolist())
        #     str_text = str_text + 'score = {} \n{}\n\n'.format(score,str(reason))
        #     str_text = str_text + '현재 시가총액 = {} 억원\n적정 시가총액 = {:.0f} 억원\n\n'.format(market_value,expectationProfit*10)
        #     str_text = str_text + '현재 주가 = {:.0f} 원\n적정 주가 = {:.0f} 원\n\n'.format(now_price,proper_price)
        #     str_text = str_text + '주당배당금 = {} 원\n'.format(' 원, '.join(annual_finance.iloc[13,0:3].tolist()))
        #     str_text = str_text + '평균 = {:.0f} 원\n배당수익률 {:.1f}%\n'.format(average_dividend, rate_dividend )
        #     str_text = str_text + '현재 배당금에 대한 목표주가 = {:.0f} 원\n평균 배당금에 대한 목표주가 = {:.0f} 원\n'.format(target_price_by_dividend, target_price_by_avg_dividend )

 
        #     my_email.sendEmail(str_subject,str_text)

    return analyzeResult

# code가 csv의 코드들중에 존재 하는지 
def isCodeExist(code, file_name):
    isExist = False
    List = pd.read_csv('{}'.format(file_name), dtype={'code': object})
    for i in List['code']:
        if code == i:
            isExist = True
            return isExist

    return isExist


#code = '005930'
#analyzeStock(code)


## main-------------------------------------------------------------------------------------

# CSV 읽어오기
file_dir = "C:/Users/조성현/Desktop/super_ant/backup/checkList/"
today = manageFile.get_today()
analyzeData = pd.read_csv(file_dir + '{}_checkList.csv'.format(today), dtype={'code': object}, thousands = ',' )
print('csv file is loaded')

print('now analyzing')

# 통과한 종목들의 엑셀 시트 만들기
file_dir = "C:/Users/조성현/Desktop/super_ant/recommend_list.xlsx"
writer = pd.ExcelWriter(file_dir, engine = 'xlsxwriter')
writer.save()

# 코드별로 반복 결과 저장
analzyeScoreList = pd.Series([])
reasonList = pd.Series([])

for stock in analyzeData['code']:

   analyzeResult = analyzeStock(stock)

   # score 이어붙이기
   analyzeScore = pd.Series(analyzeResult.iloc[0])
   analzyeScoreList = pd.concat([analzyeScoreList,analyzeScore],ignore_index=True)
   # reason 이어 붙이기
   reason = pd.Series(''.join(analyzeResult[1:].tolist()))
   reasonList = pd.concat([reasonList, reason],ignore_index= True)

# 통과한 종목들의 엑셀 시트 Sheet1 지우기
book = load_workbook(file_dir)
std = book.get_sheet_by_name('Sheet1')
book.remove_sheet(std)
book.save(file_dir)

# column 추가
analyzeData['score'] = analzyeScoreList
analyzeData['reason'] = reasonList

# 정렬
analyzeData.sort_values(by='score', ascending = False, inplace =True)

# 파일로 저장
file_dir = "C:/Users/조성현/Desktop/super_ant/backup/analysis/"
analyzeData.to_excel(file_dir + '{}_analyzeResult.xlsx'.format(today),sheet_name='Sheet1')

# 종목목록 list csv 저장
recommendList = pd.concat([analyzeData['name'], analyzeData['code'],analyzeData['score']], axis=1)
recommendList = recommendList.replace('ERROR no data','0')
print(recommendList['score'])
recommendList = recommendList[recommendList['score'] > '80' ]
recommendList.reset_index(inplace=True)
recommendList.to_csv('recommendedList.csv')